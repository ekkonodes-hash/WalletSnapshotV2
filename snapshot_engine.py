"""
snapshot_engine.py
Parallel browser screenshot engine + Excel builder.
"""

import asyncio
import uuid
import os
import io
import platform
from datetime import datetime, timezone
from pathlib import Path

from PIL import Image as PILImage, ImageDraw, ImageFont
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.drawing.image import Image as XLImage

BASE_DIR = Path(__file__).parent

# ── DATA_DIR (same logic as app.py) ──────────────────────────────────────────
DATA_DIR = Path(os.environ.get("DATA_DIR", str(BASE_DIR)))
DATA_DIR.mkdir(parents=True, exist_ok=True)

PROFILE  = DATA_DIR / "browser_profile"
OUTPUTS  = DATA_DIR / "outputs"
OUTPUTS.mkdir(exist_ok=True)
PROFILE.mkdir(exist_ok=True)

# Force headless on Linux with no display (Railway/Docker), or if HEADLESS=true is set
_on_linux_no_display = (platform.system() == "Linux" and not os.environ.get("DISPLAY"))
HEADLESS = _on_linux_no_display or os.environ.get("HEADLESS", "false").lower() == "true"

# ── Optional proxy (residential proxy or Browserbase helps bypass Cloudflare) ─
# Set PROXY_URL=http://user:pass@host:port  (or socks5://...)
PROXY_URL = os.environ.get("PROXY_URL", "")

# Realistic Chrome UA — updated periodically to match a current stable release
_CHROME_UA = (
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
    "AppleWebKit/537.36 (KHTML, like Gecko) "
    "Chrome/124.0.0.0 Safari/537.36"
)

# ── shared status dict (read by Flask for live polling) ─────────────────────
status = {
    "running"     : False,
    "message"     : "Idle — ready to run.",
    "last_run"    : None,
    "last_output" : None,
    "last_file"   : None,
    "error"       : None,
    "history"     : [],          # list of {ts, file, pages}
}

# ── helpers ──────────────────────────────────────────────────────────────────
def _fill(hex_): return PatternFill("solid", start_color=hex_, end_color=hex_)
def _border(color="CCCCCC"):
    s = Side(style="thin", color=color)
    return Border(left=s, right=s, top=s, bottom=s)

CHAIN_COLORS = {
    "TON"      : "0098EA",
    "SUI"      : "4DA2FF",
    "TRON"     : "E53935",
    "ETH"      : "627EEA",
    "SOL"      : "9945FF",
    "BTC"      : "F7931A",
    "BNB"      : "F3BA2F",
    "ALGO"     : "00BCD4",
    "SEI"      : "9C27B0",
    "STELLAR"  : "7986CB",
    "XDC"      : "1976D2",
    "BERACHAIN": "F4811F",
}
DEFAULT_CC = "6C63FF"
NAVY="1A1A2E"; WHITE="FFFFFF"; LGRAY="F5F6FA"

# ── timestamp bar stamper ─────────────────────────────────────────────────────
def _crop_height(screenshot_bytes, max_height):
    """Crop screenshot to max_height pixels (0 = no limit)."""
    if not max_height:
        return screenshot_bytes
    try:
        img = PILImage.open(io.BytesIO(screenshot_bytes))
        if img.height <= max_height:
            return screenshot_bytes
        cropped = img.crop((0, 0, img.width, max_height))
        buf = io.BytesIO()
        cropped.save(buf, format="PNG")
        return buf.getvalue()
    except Exception:
        return screenshot_bytes

def _stamp_bar(screenshot_bytes, chain, token, explorer_name, url, timestamp):
    """Burn a dark audit bar onto the bottom of a screenshot PNG."""
    try:
        img  = PILImage.open(io.BytesIO(screenshot_bytes)).convert("RGB")
        bar_h = 36
        out  = PILImage.new("RGB", (img.width, img.height + bar_h), (10, 17, 34))
        out.paste(img, (0, 0))
        draw = ImageDraw.Draw(out)

        # dark bar background
        draw.rectangle([0, img.height, img.width, img.height + bar_h],
                       fill=(10, 17, 34))
        # subtle top border line
        draw.line([0, img.height, img.width, img.height], fill=(30, 50, 80), width=1)

        # try Windows fonts first, then Linux/Docker system fonts, fall back to PIL default
        font_paths = [
            "C:/Windows/Fonts/consola.ttf",
            "C:/Windows/Fonts/cour.ttf",
            "C:/Windows/Fonts/arial.ttf",
            "/usr/share/fonts/truetype/dejavu/DejaVuSansMono.ttf",
            "/usr/share/fonts/truetype/liberation/LiberationMono-Regular.ttf",
            "/usr/share/fonts/truetype/freefont/FreeMono.ttf",
        ]
        font = None
        for fp in font_paths:
            try:
                font = ImageFont.truetype(fp, 14)
                break
            except Exception:
                pass
        if font is None:
            font = ImageFont.load_default()

        left  = f"AUDIT CAPTURE  —  {chain}  |  {explorer_name}  ·  {url}"
        right = timestamp

        y = img.height + (bar_h - 16) // 2
        draw.text((12, y), left,  fill=(52, 211, 153), font=font)   # green

        # right-align timestamp
        try:
            tw = draw.textlength(right, font=font)
        except AttributeError:
            tw = font.getlength(right)
        draw.text((img.width - tw - 12, y), right, fill=(148, 163, 184), font=font)

        buf = io.BytesIO()
        out.save(buf, format="PNG")
        return buf.getvalue()
    except Exception:
        return screenshot_bytes   # if anything fails, return original untouched


def _stamp_blocked_bar(screenshot_bytes, reason):
    """Overlay a bright red warning banner at the top of a blocked screenshot."""
    try:
        img   = PILImage.open(io.BytesIO(screenshot_bytes)).convert("RGB")
        bar_h = 48
        out   = PILImage.new("RGB", (img.width, img.height + bar_h), (180, 0, 0))
        out.paste(img, (0, bar_h))
        draw  = ImageDraw.Draw(out)
        draw.rectangle([0, 0, img.width, bar_h], fill=(180, 0, 0))

        font_paths = [
            "C:/Windows/Fonts/arialbd.ttf",
            "C:/Windows/Fonts/arial.ttf",
            "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf",
            "/usr/share/fonts/truetype/liberation/LiberationSans-Bold.ttf",
        ]
        font = None
        for fp in font_paths:
            try:
                font = ImageFont.truetype(fp, 15)
                break
            except Exception:
                pass
        if font is None:
            font = ImageFont.load_default()

        msg = f"⚠  BLOCKED — {reason}"
        draw.text((12, (bar_h - 16) // 2), msg, fill=(255, 255, 180), font=font)

        buf = io.BytesIO()
        out.save(buf, format="PNG")
        return buf.getvalue()
    except Exception:
        return screenshot_bytes


# ── bot/challenge page detection ─────────────────────────────────────────────
# Page titles / body patterns that indicate a block or challenge page.
_BLOCK_TITLES = [
    "just a moment",          # Cloudflare standard JS challenge
    "performing security verification",  # Cloudflare under-attack
    "attention required",     # Cloudflare captcha
    "access blocked",         # Routescan / generic
    "access denied",
    "403 forbidden",
    "ddos-guard",
    "enable javascript",
    "checking your browser",
    "please wait",
    "bot protection",
    "human verification",
]

async def _check_blocked(page) -> str | None:
    """Return a short description if the page looks like a bot block, else None."""
    try:
        title = (await page.title()).lower()
        for pat in _BLOCK_TITLES:
            if pat in title:
                return f"Bot challenge / block detected: \"{await page.title()}\""
    except Exception:
        pass
    return None


# ── async capture ─────────────────────────────────────────────────────────────
async def _capture(wallets, cb, wait_secs=12, max_height=3000):
    """Open explorer URLs in batches to stay within Railway's memory limits."""
    from playwright.async_api import async_playwright

    # playwright-stealth patches navigator.webdriver, plugins, languages,
    # canvas fingerprint, WebGL, chrome runtime object, and ~25 other signals
    # that Cloudflare and bot-detection services check.
    try:
        from playwright_stealth import stealth_async
        _stealth_available = True
    except ImportError:
        _stealth_available = False

    # BATCH_SIZE controls how many pages are open simultaneously.
    # Each full-page screenshot uses ~150-200 MB of RAM. Railway containers
    # have limited memory, so we process pages in small batches to avoid
    # "Target crashed" OOM errors. Set BATCH_SIZE env var to override.
    BATCH_SIZE = int(os.environ.get("BATCH_SIZE", "4"))

    tasks = []
    for w in wallets:
        addr = w.get("address", "")
        for exp in w.get("explorers", []):
            url = exp.get("url", "").replace("{address}", addr)
            if url:
                tasks.append({"wallet": w, "explorer": exp, "url": url})

    if not tasks:
        return []

    total   = len(tasks)
    batches = [tasks[i:i + BATCH_SIZE] for i in range(0, total, BATCH_SIZE)]
    cb(f"{total} pages to capture in {len(batches)} batch(es) of {BATCH_SIZE}…")
    results = []
    ts      = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC")

    async with async_playwright() as p:
        # Build proxy config if env var is set
        proxy_cfg = {"server": PROXY_URL} if PROXY_URL else None

        ctx = await p.chromium.launch_persistent_context(
            user_data_dir=str(PROFILE),
            headless=HEADLESS,
            viewport={"width": 1440, "height": 820},
            user_agent=_CHROME_UA,
            proxy=proxy_cfg,
            # Remove "--enable-automation" from Playwright's default args —
            # this flag is one of the primary signals Cloudflare looks for.
            ignore_default_args=["--enable-automation"],
            args=[
                "--disable-blink-features=AutomationControlled",
                "--disable-infobars",
                "--no-first-run",
                "--no-default-browser-check",
                "--no-sandbox",                  # required on Linux servers
                "--disable-setuid-sandbox",
                "--disable-dev-shm-usage",       # prevents crashes in Docker
                "--disable-background-timer-throttling",
                "--disable-renderer-backgrounding",
                "--js-flags=--max-old-space-size=512",  # cap V8 heap per tab
            ],
            # Pass typical browser headers so requests look organic
            extra_http_headers={
                "Accept-Language": "en-US,en;q=0.9",
                "Accept-Encoding": "gzip, deflate, br",
            },
        )

        for batch_num, batch in enumerate(batches, 1):
            cb(f"Batch {batch_num}/{len(batches)}: loading {len(batch)} page(s)…")

            # Open pages for this batch and apply stealth before navigation
            pages = []
            for t in batch:
                pg = await ctx.new_page()
                if _stealth_available:
                    await stealth_async(pg)
                pages.append((pg, t))

            # Navigate all pages in the batch simultaneously
            await asyncio.gather(
                *[pg.goto(t["url"], wait_until="domcontentloaded", timeout=45000)
                  for pg, t in pages],
                return_exceptions=True,
            )

            # Wait for dynamic content (JS-heavy explorers need this)
            cb(f"Batch {batch_num}/{len(batches)}: waiting {wait_secs}s for content…")
            await asyncio.sleep(wait_secs)

            # Scroll to top for viewport screenshots
            for pg, t in pages:
                try:
                    if not t["explorer"].get("full_page", True):
                        await pg.evaluate("window.scrollTo(0, 0);")
                        await asyncio.sleep(0.3)
                except Exception:
                    pass

            # Take screenshots one at a time — sequential avoids the memory
            # spike that causes "Target crashed" when all render simultaneously
            cb(f"Batch {batch_num}/{len(batches)}: taking screenshots…")
            for pg, t in pages:
                try:
                    # ── Bot challenge detection & retry ───────────────────
                    block_reason = await _check_blocked(pg)
                    if block_reason:
                        # Cloudflare JS challenges can pass with extra wait.
                        # Give it 20 more seconds then check again.
                        cb(f"Challenge detected on {t['explorer'].get('name','')} — waiting 20s for JS challenge…")
                        await asyncio.sleep(20)
                        block_reason = await _check_blocked(pg)

                    ss = await pg.screenshot(
                        full_page=t["explorer"].get("full_page", True),
                        timeout=60000,   # 60s — heavy explorers need this
                    )
                    if t["explorer"].get("full_page", True) and max_height:
                        ss = _crop_height(ss, max_height)

                    # If still blocked after retry, stamp a red warning bar
                    # so it's unmistakably visible in the Excel report
                    if block_reason:
                        ss = _stamp_blocked_bar(ss, block_reason)

                    stamped = _stamp_bar(
                        ss,
                        chain         = t["wallet"].get("chain", ""),
                        token         = t["wallet"].get("token", ""),
                        explorer_name = t["explorer"].get("name", ""),
                        url           = t["url"],
                        timestamp     = ts,
                    )
                    results.append({
                        "wallet"    : t["wallet"],
                        "explorer"  : t["explorer"],
                        "url"       : t["url"],
                        "timestamp" : ts,
                        "screenshot": stamped,
                        # Surface the block reason as a soft error so the
                        # Summary tab shows ✗ and the reason is visible
                        "error"     : block_reason,
                    })
                except Exception as e:
                    results.append({
                        "wallet"    : t["wallet"],
                        "explorer"  : t["explorer"],
                        "url"       : t["url"],
                        "timestamp" : ts,
                        "screenshot": None,
                        "error"     : str(e),
                    })

            # Close batch pages to free memory before the next batch
            for pg, _ in pages:
                try:
                    await pg.close()
                except Exception:
                    pass

        await ctx.close()

    return results


# ── Excel builder ─────────────────────────────────────────────────────────────
def _build_excel(results, ts_str):
    wb = Workbook()
    wb.remove(wb.active)

    # group by chain
    chains = {}
    for r in results:
        c = r["wallet"].get("chain", "Other")
        chains.setdefault(c, []).append(r)

    # ── Summary ──────────────────────────────────────────────────────────────
    ws = wb.create_sheet("Summary")
    ws.sheet_view.showGridLines = False
    for col, w in zip("ABCDEFG", [4,12,18,48,16,10,26]):
        ws.column_dimensions[col].width = w

    ws.merge_cells("A1:G1")
    c = ws["A1"]
    c.value  = "WALLET BALANCE SNAPSHOT"
    c.font   = Font(name="Arial", bold=True, size=16, color=WHITE)
    c.fill   = _fill(NAVY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 38

    ws.merge_cells("A2:G2")
    c = ws["A2"]
    c.value = (f"Captured: {ts_str}   |   "
               f"{len(results)} pages   |   {len(chains)} chain(s)")
    c.font  = Font(name="Arial", size=10, color="8892B0", italic=True)
    c.fill  = _fill(NAVY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 20
    ws.merge_cells("A3:G3"); ws["A3"].fill = _fill(NAVY)
    ws.row_dimensions[3].height = 6

    for i, h in enumerate(["","Chain","Explorer","Address","Token","OK?","Captured"],1):
        c = ws.cell(row=4, column=i, value=h)
        c.font  = Font(name="Arial", bold=True, size=10, color=WHITE)
        c.fill  = _fill("2D3561")
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = _border(WHITE)
    ws.row_dimensions[4].height = 20

    for ri, r in enumerate(results, 5):
        chain = r["wallet"].get("chain","?")
        cc    = CHAIN_COLORS.get(chain, DEFAULT_CC)
        bg    = LGRAY if ri % 2 == 0 else WHITE
        ws.cell(row=ri, column=1).fill   = _fill(cc)
        ws.cell(row=ri, column=1).border = _border()
        ok = "✓" if (r["screenshot"] and not r.get("error")) else "✗"
        for ci, v in enumerate([chain,
                                 r["explorer"].get("name",""),
                                 r["wallet"].get("address",""),
                                 r["wallet"].get("token",""),
                                 ok, r["timestamp"]], 2):
            c = ws.cell(row=ri, column=ci, value=v)
            c.font = Font(name="Arial", size=10,
                          color="00873E" if v=="✓" else ("CC0000" if v=="✗" else "333333"))
            c.fill = _fill(bg)
            c.alignment = Alignment(horizontal="left", vertical="center")
            c.border = _border()
        ws.row_dimensions[ri].height = 18

    # ── One tab per chain ─────────────────────────────────────────────────────
    tmp_files = []   # collect temp PNGs; delete AFTER wb.save()
    for chain_name, chain_results in chains.items():
        cc = CHAIN_COLORS.get(chain_name, DEFAULT_CC)
        ws = wb.create_sheet(chain_name)
        ws.sheet_view.showGridLines = False
        for col, w in zip("ABCDEF", [3,22,52,18,18,3]):
            ws.column_dimensions[col].width = w

        ws.merge_cells("A1:F1")
        c = ws["A1"]
        c.value = f"  {chain_name}  —  Snapshot   |   {ts_str}"
        c.font  = Font(name="Arial", bold=True, size=14, color=WHITE)
        c.fill  = _fill(cc)
        c.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[1].height = 34

        row = 3
        for r in chain_results:
            w_info = r["wallet"]; exp = r["explorer"]

            # explorer header bar
            ws.merge_cells(start_row=row,start_column=2,end_row=row,end_column=5)
            c = ws.cell(row=row, column=2,
                        value=f"  {exp.get('name','')}   ·   {r['url']}")
            c.font  = Font(name="Arial", bold=True, size=11, color=WHITE)
            c.fill  = _fill(cc)
            c.alignment = Alignment(horizontal="left", vertical="center")
            ws.row_dimensions[row].height = 24; row += 1

            # metadata rows
            for label, val in [
                ("Address" , w_info.get("address","")),
                ("Token"   , w_info.get("token","")),
                ("Captured", r["timestamp"]),
            ]:
                cl = ws.cell(row=row, column=2, value=label)
                cl.font = Font(name="Arial", bold=True, size=10, color="555555")
                cl.fill = _fill(LGRAY); cl.border = _border()
                cl.alignment = Alignment(horizontal="left", vertical="center")
                ws.merge_cells(start_row=row,start_column=3,end_row=row,end_column=5)
                cv = ws.cell(row=row, column=3, value=val)
                cv.font = Font(name="Arial", size=10, color=NAVY)
                cv.fill = _fill(WHITE); cv.border = _border()
                cv.alignment = Alignment(horizontal="left", vertical="center")
                ws.row_dimensions[row].height = 18; row += 1

            row += 1   # spacer

            # embed screenshot
            if r["screenshot"]:
                tmp = BASE_DIR / f"_tmp_{uuid.uuid4().hex}.png"
                tmp.write_bytes(r["screenshot"])
                tmp_files.append(tmp)          # delete after save, not now
                img = XLImage(str(tmp))
                # maintain aspect ratio — fit to 900px wide, scale height proportionally
                raw = PILImage.open(io.BytesIO(r["screenshot"]))
                ratio = raw.height / raw.width if raw.width else 1
                img.width  = 900
                img.height = int(900 * ratio)
                ws.add_image(img, f"B{row}")
                # each Excel row ≈ 20pt tall; calculate how many rows the image needs
                n_rows = max(1, int(img.height / 20) + 1)
                for r2 in range(row, row + n_rows):
                    ws.row_dimensions[r2].height = 20
                row += n_rows + 1
            else:
                c = ws.cell(row=row, column=2,
                            value=f"⚠  Screenshot failed: {r.get('error','')}")
                c.font = Font(name="Arial", size=10, color="CC0000")
                row += 2

            row += 2   # gap between explorers

    # save — temp PNGs must still exist at this point
    safe_ts = ts_str.replace(" ","_").replace(":","").replace("/","-")
    out = OUTPUTS / f"Snapshot_{safe_ts}.xlsx"
    wb.save(str(out))

    # now safe to delete temp files
    for tmp in tmp_files:
        try: tmp.unlink()
        except: pass

    return str(out)


# ── public synchronous entry point ────────────────────────────────────────────
def run_snapshot(wallets, wait_secs=12, max_height=3000):
    """Called from a background thread by Flask."""
    global status
    status.update(running=True, error=None, message="Starting…")

    def cb(msg):
        status["message"] = msg

    try:
        loop = asyncio.new_event_loop()
        asyncio.set_event_loop(loop)
        results  = loop.run_until_complete(_capture(wallets, cb, wait_secs=wait_secs, max_height=max_height))
        loop.close()

        cb("Building Excel report…")
        ts_str   = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC")
        out_path = _build_excel(results, ts_str)
        fname    = Path(out_path).name

        status.update(
            running=False, last_run=ts_str,
            last_output=out_path, last_file=fname,
            message=f"Done ✓  →  {fname}",
        )
        status["history"].insert(0, {"ts": ts_str, "file": fname,
                                     "pages": len(results)})
        status["history"] = status["history"][:20]   # keep last 20
        return out_path

    except Exception as e:
        status.update(running=False, error=str(e),
                      message=f"Error: {e}")
        raise
